#!/usr/bin/env python3
import argparse
import csv
from pathlib import Path

from openpyxl import Workbook


EXCEL_MAX_ROWS = 1048576


def value(v):
    if v == "NA":
        return v
    try:
        if "." in v:
            return float(v)
        return int(v)
    except ValueError:
        return v


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--summary-tsv", type=Path, required=True)
    parser.add_argument("--technical-tsv", type=Path, required=True)
    parser.add_argument("--summary-xlsx", type=Path, required=True)
    parser.add_argument("--technical-xlsx", type=Path, required=True)
    args = parser.parse_args()

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    with args.summary_tsv.open(newline="", encoding="utf-8") as handle:
        for row in csv.reader(handle, delimiter="\t"):
            ws.append([value(x) for x in row])
    wb.save(args.summary_xlsx)

    wb = Workbook(write_only=True)
    sheet = None
    row_count = EXCEL_MAX_ROWS
    with args.technical_tsv.open(newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle, delimiter="\t")
        header = next(reader)
        for row in reader:
            if row_count >= EXCEL_MAX_ROWS:
                sheet = wb.create_sheet("hits_{}".format(len(wb.worksheets) + 1))
                sheet.append(header)
                row_count = 1
            sheet.append([value(x) for x in row])
            row_count += 1
    if sheet is None:
        sheet = wb.create_sheet("hits_1")
        sheet.append(header)
    wb.save(args.technical_xlsx)


if __name__ == "__main__":
    main()
