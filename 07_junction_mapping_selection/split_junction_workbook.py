#!/usr/bin/env python3
import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


def backbone_bounds(value):
    match = re.fullmatch(r"(\d+)_(\d+)", str(value))
    if not match:
        raise ValueError(f"Invalid backbone interval: {value}")
    return int(match.group(1)), int(match.group(2))


def main():
    if len(sys.argv) != 3:
        raise SystemExit(f"usage: {sys.argv[0]} INPUT.xlsx OUTPUT.xlsx")
    source_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    workbook = load_workbook(source_path)
    source = workbook.active
    values = list(source.iter_rows(values_only=True))
    header = values[0]
    rows = values[1:]
    d_index = header.index("backbone_on_reference")
    g_index = header.index("gap_length")

    near_5042 = []
    gap0_far = []
    remaining = []
    for row in rows:
        first, second = backbone_bounds(row[d_index])
        is_near = abs(first - 5042) <= 10 or abs(second - 5042) <= 10
        if is_near:
            near_5042.append(row)
        elif int(row[g_index]) == 0:
            gap0_far.append(row)
        else:
            remaining.append(row)

    workbook.remove(source)
    groups = [
        ("near_5042", near_5042),
        ("gap0_far", gap0_far),
        ("remaining", remaining),
    ]
    for name, group in groups:
        sheet = workbook.create_sheet(name)
        sheet.append(header)
        for row in group:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = copy(cell.font)
        for col, width in {"A": 38, "B": 14, "C": 20, "D": 25, "E": 18, "F": 22, "G": 25, "H": 14, "I": 12, "J": 18, "K": 14, "L": 12}.items():
            sheet.column_dimensions[col].width = width

    workbook.save(output_path)
    print(f"near_5042\t{len(near_5042)}")
    print(f"gap0_far\t{len(gap0_far)}")
    print(f"remaining\t{len(remaining)}")
    print(f"total\t{len(near_5042) + len(gap0_far) + len(remaining)}")


if __name__ == "__main__":
    main()
