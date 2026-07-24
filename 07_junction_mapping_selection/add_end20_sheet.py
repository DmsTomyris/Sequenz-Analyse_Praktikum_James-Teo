#!/usr/bin/env python3
import sys
from copy import copy
from openpyxl import load_workbook


def interval(value):
    start, end = str(value).split("_")
    return int(start), int(end)


def main():
    if len(sys.argv) != 3:
        raise SystemExit(f"usage: {sys.argv[0]} INPUT.xlsx OUTPUT.xlsx")
    input_path, output_path = sys.argv[1:]
    workbook = load_workbook(input_path)
    rows = []
    header = None
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            continue
        if header is None:
            header = list(values[0])
        indices = {name: i for i, name in enumerate(values[0])}
        for row in values[1:]:
            if not row:
                continue
            backbone_start, backbone_end = interval(row[indices["backbone_on_reference"]])
            insert_start, insert_end = interval(row[indices["insert_on_reference"]])
            insert_length = int(row[indices["insert_length"]])
            backbone_near = min(abs(backbone_start - 5042), abs(backbone_end - 5042)) <= 20
            insert_near = min(insert_start - 1, insert_length - insert_end) <= 20
            if backbone_near and insert_near:
                rows.append(row)

    if "end20_both" in workbook.sheetnames:
        del workbook["end20_both"]
    sheet = workbook.create_sheet("end20_both")
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = copy(cell.font)
    for col, width in {"A": 38, "B": 14, "C": 20, "D": 25, "E": 18, "F": 22, "G": 25, "H": 14, "I": 12, "J": 18, "K": 14, "L": 12}.items():
        sheet.column_dimensions[col].width = width
    workbook.save(output_path)
    print(f"end20_both\t{len(rows)}")


if __name__ == "__main__":
    main()
