from collections import Counter
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE = Path(__file__).resolve().parent
CURRENT = BASE / "junction_mappings_selected_hits.xlsx"
BEFORE_MAPQ_SELECTION = BASE / "junction_mappings_corrected_grouped_mapq.xlsx"
OUTPUT = BASE / "junction_mappings_selected_hits_20bp.xlsx"
TARGET = 5042
LIMIT = 20


def parse_interval(value):
    start, end = value.split("_")
    return int(start), int(end)


def distance_to_backbone_target(value):
    start, end = parse_interval(value)
    if start <= TARGET <= end:
        return 0
    return min(abs(TARGET - start), abs(TARGET - end))


def distance_to_insert_end(row, indexes):
    start, end = parse_interval(row[indexes["insert_on_reference"]])
    insert_length = int(row[indexes["insert_length"]])
    return min(start - 1, insert_length - end)


def qualifies(row, indexes):
    return (
        distance_to_backbone_target(row[indexes["backbone_on_reference"]]) <= LIMIT
        and distance_to_insert_end(row, indexes) <= LIMIT
    )


def copy_cell(source, target):
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def read_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    first = workbook.worksheets[0]
    headers = [cell.value for cell in next(first.iter_rows())]
    indexes = {name: i for i, name in enumerate(headers)}
    rows = []
    for sheet in workbook.worksheets:
        rows.extend(list(sheet.iter_rows(min_row=2, values_only=True)))
    workbook.close()
    return headers, indexes, rows


def main():
    source_wb = load_workbook(CURRENT)
    source_sheet = source_wb.worksheets[0]
    headers = [cell.value for cell in source_sheet[1]]
    indexes = {name: i for i, name in enumerate(headers)}
    all_rows = []
    for sheet in source_wb.worksheets:
        all_rows.extend(list(sheet.iter_rows(min_row=2, values_only=True)))

    qualifying = [row for row in all_rows if qualifies(row, indexes)]
    nonqualifying = [row for row in all_rows if not qualifies(row, indexes)]

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    output_sheets = {name: out_wb.create_sheet(name) for name in ["near_5042", "gap0_far", "remaining", "within_20bp"]}
    for name, out_sheet in output_sheets.items():
        for col, value in enumerate(headers, 1):
            target = out_sheet.cell(1, col, value)
            copy_cell(source_sheet.cell(1, col), target)
        out_sheet.freeze_panes = "A2"
        for letter, dimension in source_sheet.column_dimensions.items():
            out_sheet.column_dimensions[letter].width = dimension.width

    # Preserve the three existing non-hit-count groups, removing qualifying rows.
    for source_sheet_existing in source_wb.worksheets:
        target_sheet = output_sheets[source_sheet_existing.title]
        for values in source_sheet_existing.iter_rows(min_row=2, values_only=True):
            if qualifies(values, indexes):
                continue
            target_row = target_sheet.max_row + 1
            for col, value in enumerate(values, 1):
                target = target_sheet.cell(target_row, col, value)
                copy_cell(source_sheet_existing.cell(target_row if target_row <= source_sheet_existing.max_row else 2, col), target)

    # Put all qualifying rows into the fourth sheet.
    target_sheet = output_sheets["within_20bp"]
    for values in qualifying:
        target_row = target_sheet.max_row + 1
        for col, value in enumerate(values, 1):
            target = target_sheet.cell(target_row, col, value)
            copy_cell(source_sheet.cell(2, col), target)

    for sheet in output_sheets.values():
        last_row = sheet.max_row
        last_col = len(headers)
        sheet.auto_filter.ref = f"A1:{chr(64 + last_col)}{last_row}"

    out_wb.save(OUTPUT)

    old_headers, old_indexes, old_rows = read_rows(BEFORE_MAPQ_SELECTION)
    old_qualifying = [row for row in old_rows if qualifies(row, old_indexes)]
    print(f"current_total_pairs={len(all_rows)}")
    print(f"current_qualifying_pairs={len(qualifying)}")
    print(f"current_qualifying_reads={len({row[indexes['read_id']] for row in qualifying})}")
    print(f"before_mapq_selection_total_pairs={len(old_rows)}")
    print(f"before_mapq_selection_qualifying_pairs={len(old_qualifying)}")
    print(f"before_mapq_selection_qualifying_reads={len({row[old_indexes['read_id']] for row in old_qualifying})}")
    print(f"removed_from_first_three={len(qualifying)}")
    print(f"output={OUTPUT}")


if __name__ == "__main__":
    main()
