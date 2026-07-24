from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import load_workbook, Workbook


BASE = Path(__file__).resolve().parent
INPUT = BASE / "junction_mappings_corrected_grouped_mapq.xlsx"
OUTPUT = BASE / "junction_mappings_selected_hits.xlsx"
BACKBONE_TARGET = 5042


def interval(value):
    start, end = value.split("_")
    return int(start), int(end)


def distance_to_interval(position, value):
    start, end = interval(value)
    if start <= position <= end:
        return 0
    return min(abs(position - start), abs(position - end))


def insert_end_distance(row, indexes):
    start, end = interval(row[indexes["insert_on_reference"]])
    insert_length = int(row[indexes["insert_length"]])
    return min(start - 1, insert_length - end)


def gap_values(first, second):
    if first[1] < second[0]:
        gap = second[0] - first[1] - 1
    elif second[1] < first[0]:
        gap = first[0] - second[1] - 1
    else:
        gap = 0
    if first[1] < second[0]:
        signed = gap
    elif second[1] < first[0]:
        signed = gap
    else:
        signed = -(min(first[1], second[1]) - max(first[0], second[0]) + 1)
    return gap, signed


def copy_cell(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def main():
    source_wb = load_workbook(INPUT)
    source_sheet = source_wb.worksheets[0]
    headers = [cell.value for cell in source_sheet[1]]
    indexes = {name: i for i, name in enumerate(headers)}
    required = [
        "read_id", "backbone_on_reference", "insert_on_reference",
        "insert_length", "backbone_mapq", "insert_mapq", "gap_length",
    ]
    if any(name not in indexes for name in required):
        raise ValueError("Input workbook does not have the expected no-hit-count columns")

    all_rows = []
    by_read = defaultdict(list)
    for sheet in source_wb.worksheets:
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = list(values)
            all_rows.append(row)
            by_read[row[indexes["read_id"]]].append(row)

    selected = []
    for read_id, rows in sorted(by_read.items()):
        backbone_candidates = {}
        insert_candidates = {}
        for row in rows:
            backbone_key = (
                row[indexes["backbone_on_reference"]],
                row[indexes["read_on_backbone"]],
                row[indexes["backbone_mapq"]],
            )
            insert_key = (
                row[indexes["insert_name"]],
                row[indexes["insert_on_reference"]],
                row[indexes["read_on_insert"]],
                row[indexes["insert_mapq"]],
            )
            backbone_candidates[backbone_key] = row
            insert_candidates[insert_key] = row

        best_backbone = min(
            backbone_candidates.values(),
            key=lambda row: (
                distance_to_interval(BACKBONE_TARGET, row[indexes["backbone_on_reference"]]),
                -int(row[indexes["backbone_mapq"]]),
                -(
                    interval(row[indexes["backbone_on_reference"]])[1]
                    - interval(row[indexes["backbone_on_reference"]])[0]
                    + 1
                ),
                row[indexes["backbone_on_reference"]],
            ),
        )
        best_insert = min(
            insert_candidates.values(),
            key=lambda row: (
                insert_end_distance(row, indexes),
                -int(row[indexes["insert_mapq"]]),
                row[indexes["insert_name"]],
                row[indexes["insert_on_reference"]],
            ),
        )

        combined = list(best_backbone)
        for name in ["read_on_insert", "insert_name", "insert_on_reference", "insert_length", "insert_mapq"]:
            combined[indexes[name]] = best_insert[indexes[name]]
        gap, signed_gap = gap_values(
            interval(best_backbone[indexes["read_on_backbone"]]),
            interval(best_insert[indexes["read_on_insert"]]),
        )
        combined[indexes["gap_length"]] = gap
        combined[indexes["gap_length_signed"]] = "+" if signed_gap > 0 else str(signed_gap)
        selected.append(combined)

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    sheets = {name: out_wb.create_sheet(name) for name in ["near_5042", "gap0_far", "remaining"]}
    for name, out_sheet in sheets.items():
        for col, value in enumerate(headers, 1):
            target = out_sheet.cell(1, col, value)
            copy_cell(source_sheet.cell(1, col), target)
        out_sheet.freeze_panes = "A2"
        out_sheet.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"
        for letter, dimension in source_sheet.column_dimensions.items():
            out_sheet.column_dimensions[letter].width = dimension.width

    counts = defaultdict(int)
    for row in selected:
        backbone_distance = distance_to_interval(BACKBONE_TARGET, row[indexes["backbone_on_reference"]])
        if backbone_distance <= 10:
            sheet_name = "near_5042"
        elif int(row[indexes["gap_length"]]) == 0:
            sheet_name = "gap0_far"
        else:
            sheet_name = "remaining"
        out_sheet = sheets[sheet_name]
        target_row = out_sheet.max_row + 1
        for col, value in enumerate(row, 1):
            target = out_sheet.cell(target_row, col, value)
            copy_cell(source_sheet.cell(2, col), target)
        counts[sheet_name] += 1
        out_sheet.auto_filter.ref = f"A1:{chr(64 + len(headers))}{target_row}"

    out_wb.save(OUTPUT)
    print(f"selected_reads={len(selected)}")
    print(f"sheet_counts={dict(counts)}")
    print(f"output={OUTPUT}")


if __name__ == "__main__":
    main()
