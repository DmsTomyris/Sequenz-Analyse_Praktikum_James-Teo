import gzip
import re
import struct
from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
BAM = ROOT / "Day2" / "Pipeline2" / "targeted_reference.map-ont.bam"
FASTQ = ROOT / "Day2" / "Pipeline2" / "dorado_reads.fastq"
INPUT_XLSX = Path(__file__).resolve().parent / "junction_mappings_corrected_grouped_mapq.xlsx"
OUTPUT_XLSX = Path(__file__).resolve().parent / "junction_mappings_corrected_grouped_mapq_with_hit_counts.xlsx"
TOLERANCE = 10
CIGAR_OPS = "MIDNSHP=X"
QUERY_OPS = set("MIS=XH")
REF_OPS = set("MDN=X")


def read_bam(path):
    with gzip.open(path, "rb") as handle:
        magic = handle.read(4)
        if magic != b"BAM\1":
            raise ValueError("Not a BAM file")
        header_len = struct.unpack("<i", handle.read(4))[0]
        handle.read(header_len)
        n_ref = struct.unpack("<i", handle.read(4))[0]
        refs = []
        for _ in range(n_ref):
            name_len = struct.unpack("<i", handle.read(4))[0]
            name = handle.read(name_len).rstrip(b"\0").decode()
            handle.read(4)
            refs.append(name)

        read_lengths = read_fastq_lengths(FASTQ)
        records = defaultdict(lambda: {"backbone": [], "insert": []})
        while True:
            block = handle.read(4)
            if not block:
                break
            block_size = struct.unpack("<i", block)[0]
            payload = handle.read(block_size)
            if len(payload) != block_size:
                raise ValueError("Truncated BAM alignment block")
            ref_id, pos, bin_mq_nl, flag_nc, read_len, _, _, _ = struct.unpack(
                "<iiIIiiii", payload[:32]
            )
            if ref_id < 0:
                continue
            name_len = bin_mq_nl & 0xFF
            mapq = (bin_mq_nl >> 8) & 0xFF
            flag = flag_nc >> 16
            n_cigar = flag_nc & 0xFFFF
            offset = 32
            read_id = payload[offset : offset + name_len - 1].decode()
            offset += name_len
            cigar_values = struct.unpack(f"<{n_cigar}I", payload[offset : offset + 4 * n_cigar])
            offset += 4 * n_cigar
            cigar = [(value >> 4, CIGAR_OPS[value & 0xF]) for value in cigar_values]
            qpos = 0
            rpos = pos + 1
            qstart = qend = rstart = rend = None
            for length, op in cigar:
                if op in QUERY_OPS:
                    if op in "M=X":
                        if qstart is None:
                            qstart = qpos + 1
                        qend = qpos + length
                    qpos += length
                if op in REF_OPS:
                    if op in "M=X":
                        if rstart is None:
                            rstart = rpos
                        rend = rpos + length - 1
                    rpos += length
            if qstart is None or rstart is None:
                continue
            full_read_len = read_lengths.get(read_id, read_len)
            if flag & 16:
                qstart, qend = full_read_len - qend + 1, full_read_len - qstart + 1
            ref_name = refs[ref_id]
            category = "backbone" if ref_name == "pGP564" else "insert"
            records[read_id][category].append(
                {
                    "ref_name": ref_name,
                    "qstart": qstart,
                    "qend": qend,
                    "rstart": rstart,
                    "rend": rend,
                    "mapq": mapq,
                    "flag": flag,
                }
            )
    return records


def read_fastq_lengths(path):
    lengths = {}
    with path.open(encoding="utf-8", errors="replace") as handle:
        while True:
            header = handle.readline()
            if not header:
                break
            sequence = handle.readline().strip()
            handle.readline()
            handle.readline()
            lengths[header[1:].split()[0]] = len(sequence)
    return lengths


def parse_interval(value):
    start, end = value.split("_")
    return int(start), int(end)


def find_record(records, ref_name, qinterval, rinterval, mapq):
    candidates = [
        record
        for record in records
        if (
            record["ref_name"] == ref_name
            and (record["qstart"], record["qend"]) == qinterval
            and (record["rstart"], record["rend"]) == rinterval
            and record["mapq"] == mapq
        )
    ]
    return candidates[0] if candidates else None


def count_other_good(records, selected):
    if selected is None:
        return None
    threshold = max(0, selected["mapq"] - TOLERANCE)
    return sum(1 for record in records if record is not selected and record["mapq"] >= threshold)


def copy_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def main():
    records = read_bam(BAM)
    workbook = load_workbook(INPUT_XLSX)
    total_rows = 0
    missing = 0
    for sheet in workbook.worksheets:
        headers = [cell.value for cell in sheet[1]]
        required = ["read_id", "read_on_backbone", "backbone_on_reference", "insert_name", "insert_on_reference", "backbone_mapq", "insert_mapq"]
        if any(name not in headers for name in required):
            raise ValueError(f"Missing headers in {sheet.title}")
        # Insert after backbone_mapq, then after insert_mapq.
        backbone_col = headers.index("backbone_mapq") + 1
        sheet.insert_cols(backbone_col + 1)
        sheet.cell(1, backbone_col + 1).value = "additional_backbone_hits_mapq10"
        headers.insert(backbone_col, "additional_backbone_hits_mapq10")
        insert_col = headers.index("insert_mapq") + 1
        sheet.insert_cols(insert_col + 1)
        sheet.cell(1, insert_col + 1).value = "additional_insert_hits_mapq10"
        headers.insert(insert_col, "additional_insert_hits_mapq10")
        copy_style(sheet.cell(1, backbone_col), sheet.cell(1, backbone_col + 1))
        copy_style(sheet.cell(1, insert_col), sheet.cell(1, insert_col + 1))
        for row in range(2, sheet.max_row + 1):
            total_rows += 1
            read_id = sheet.cell(row, headers.index("read_id") + 1).value
            backbone_name = "pGP564"
            backbone_q = parse_interval(sheet.cell(row, headers.index("read_on_backbone") + 1).value)
            backbone_r = parse_interval(sheet.cell(row, headers.index("backbone_on_reference") + 1).value)
            backbone_mapq = int(sheet.cell(row, headers.index("backbone_mapq") + 1).value)
            insert_name = sheet.cell(row, headers.index("insert_name") + 1).value
            insert_q = parse_interval(sheet.cell(row, headers.index("read_on_insert") + 1).value)
            insert_r = parse_interval(sheet.cell(row, headers.index("insert_on_reference") + 1).value)
            insert_mapq = int(sheet.cell(row, headers.index("insert_mapq") + 1).value)
            read_records = records.get(read_id, {"backbone": [], "insert": []})
            backbone = find_record(read_records["backbone"], backbone_name, backbone_q, backbone_r, backbone_mapq)
            insert = find_record(read_records["insert"], insert_name, insert_q, insert_r, insert_mapq)
            if backbone is None or insert is None:
                missing += 1
            sheet.cell(row, headers.index("additional_backbone_hits_mapq10") + 1).value = count_other_good(read_records["backbone"], backbone)
            sheet.cell(row, headers.index("additional_insert_hits_mapq10") + 1).value = count_other_good(read_records["insert"], insert)
        for row in range(2, sheet.max_row + 1):
            copy_style(sheet.cell(row, backbone_col), sheet.cell(row, backbone_col + 1))
            copy_style(sheet.cell(row, insert_col), sheet.cell(row, insert_col + 1))
            sheet.cell(row, backbone_col + 1).number_format = "0"
            sheet.cell(row, insert_col + 1).number_format = "0"
        sheet.column_dimensions["M"].width = 28
        sheet.column_dimensions["O"].width = 26
    workbook.save(OUTPUT_XLSX)
    print(f"rows={total_rows}")
    print(f"missing_alignment_matches={missing}")
    print(f"output={OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
